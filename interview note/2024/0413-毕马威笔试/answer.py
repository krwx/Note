import json

import requests

from openpyxl import Workbook, load_workbook

from selenium import webdriver
from PIL import Image
import io

from datetime import datetime
import time

import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.header import Header
from email.mime.application import MIMEApplication


# 接口的URL
url = 'https://hot.cigh.cn/weibo'


def send_request():
    # 发送GET请求
    response = requests.get(url)

    # 检查响应状态
    if response.status_code == 200:
        # 获取响应内容（假设是JSON格式）
        data = response.json()
        write_raw_info(data)
        get_top_10_info(data)
        send_email()
    else:
        print('Failed to retrieve data: ', response.status_code)


# 写入原始返回值
def write_raw_info(data):
    with open('RawInfo.txt', 'w', encoding='utf-8') as file:
        file.write(json.dumps(data, ensure_ascii=False))


# 获取前十条视频的信息
def get_top_10_info(data):
    video_list = data['data']
    video_list.sort(key=lambda ele: ele.get('hot', 0), reverse=True)

    # 创建一个新的工作簿
    wb = Workbook()
    ws = wb.active
    ws['A1'] = '标题'
    ws['B1'] = 'url'
    ws['C1'] = '热度'

    top_ten_video = video_list[:10]

    for x in top_ten_video:
        ws.append((x['title'], x['url'], x['hot']))

    # 保存工作簿到文件
    wb.save('Top10.xlsx')
    print_screen(top_ten_video)


# 截图
def print_screen(video_list):
    # 打开Chrome浏览器
    driver = webdriver.Chrome()

    # 最大化窗口，以便捕获整个网页
    driver.maximize_window()

    screen_info = []

    for video in video_list:
        driver.get(video['url'])
        time.sleep(3)

        # 截图并保存
        screenshot = driver.get_screenshot_as_png()
        # 将PNG格式的截图转换为JPG格式
        image = Image.open(io.BytesIO(screenshot))
        # 获取当前时间
        now = get_now_time_str()
        file_name = now+'.jpg'
        image.convert('RGB').save('screenshot/'+file_name, 'JPEG', quality=90)

        screen_info.append((file_name, now))

    # 清理：关闭浏览器窗口
    driver.quit()

    save_screen_info(screen_info)


# 保存截图信息
def save_screen_info(screen_info):
    workbook = load_workbook('Top10.xlsx')
    sheet = workbook.active
    sheet['D1'] = '截图文件名'
    sheet['E1'] = '截图时间'

    for index, info in enumerate(screen_info):
        temp_index = str(index+2)
        sheet['D'+temp_index] = info[0]
        sheet['E'+temp_index] = info[1]

    workbook.save('Top10.xlsx')


def get_now_time_str():
    now = datetime.now()
    # 将日期和时间格式化为 "yyyyMMddHHmmss" 格式
    formatted_time = now.strftime("%Y%m%d%H%M%S")
    return formatted_time


def send_email():
    # 创建邮件对象
    msg = MIMEMultipart()
    # 发件人的昵称
    # msg['From'] = Header('krwxkar@163.com', 'utf-8')  # 其他邮箱的写法
    msg['From'] = Header('krwx <krwxkar@163.com>')  # qq邮箱的写法
    msg['To'] = Header('625454036@qq.com', 'utf-8')  # 收件人的昵称
    msg['Subject'] = Header('请查收当日推荐', 'utf-8')  # 定义主题内容

    # 添加邮件正文
    mail_content = "您好，附件是当日推荐信息，请查收，谢谢。"
    msg.attach(MIMEText(mail_content, 'plain', 'utf-8'))

    # 添加附件
    with open('RawInfo.txt', 'rb') as file:
        attachment = MIMEApplication(file.read(), Name='RawInfo.txt')
        attachment['Content-Disposition'] = f'attachment; filename=RawInfo.txt'
        msg.attach(attachment)
    with open('Top10.xlsx', 'rb') as file:
        attachment = MIMEApplication(file.read(), Name='Top10.xlsx')
        attachment['Content-Disposition'] = f'attachment; filename=Top10.xlsx'
        msg.attach(attachment)

    # 连接到SMTP服务器
    server = smtplib.SMTP('smtp.163.com', 25)  # 使用您的SMTP服务器和端口
    server.starttls()  # 启用TLS
    server.login('krwxkar@163.com', '1372weee')  # 登录到邮件服务器

    # 发送邮件
    server.sendmail('krwxkar@163.com', '625454036@qq.com', msg.as_string())

    # 关闭连接
    server.quit()

    # 保存邮件内容到文件
    with open('Email.msg', 'w', encoding='utf-8') as file:
        file.write(msg.as_string())


if __name__ == '__main__':
    send_request()
